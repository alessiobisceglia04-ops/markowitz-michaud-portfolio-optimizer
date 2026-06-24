import pandas as pd
import numpy as np
from scipy.optimize import minimize
import openpyxl


# =============================================================================
# FASE 2 — MARKOWITZ
# =============================================================================

def calcola_matrice_covarianza(rendimenti_storici_df):
    solo_rendimenti = rendimenti_storici_df.select_dtypes(include=[np.number])
    return solo_rendimenti.cov().values


def calcola_varianza_portafoglio(pesi, matrice_cov):
    return pesi @ matrice_cov @ pesi


def minimizza_varianza(matrice_cov, vincoli_pesi, rendimento_target=None, rendimenti_attesi=None):
    """
    Minimizza la varianza del portafoglio rispettando i vincoli di peso.
    Questa versione NON blocca il programma se scipy segnala un problema:
    stampa un warning e continua, come faceva di fatto il tuo codice originale.
    """
    varianze_individuali = np.diag(matrice_cov)

    # Protezione base per evitare divisioni strane
    varianze_individuali = np.where(varianze_individuali <= 0, 1e-8, varianze_individuali)

    pesi_inv_varianza = 1.0 / varianze_individuali
    pesi_iniziali = pesi_inv_varianza / pesi_inv_varianza.sum()

    lower_bounds = np.array([b[0] for b in vincoli_pesi])
    upper_bounds = np.array([b[1] for b in vincoli_pesi])

    pesi_iniziali = np.clip(pesi_iniziali, lower_bounds, upper_bounds)

    if pesi_iniziali.sum() == 0:
        pesi_iniziali = np.ones(len(vincoli_pesi)) / len(vincoli_pesi)
    else:
        pesi_iniziali = pesi_iniziali / pesi_iniziali.sum()

    lista_vincoli = [
        {"type": "eq", "fun": lambda w: np.sum(w) - 1}
    ]

    if rendimento_target is not None:
        lista_vincoli.append({
            "type": "eq",
            "fun": lambda w, r=rendimento_target: w @ rendimenti_attesi - r,
        })

    risultato = minimize(
        fun=lambda w: calcola_varianza_portafoglio(w, matrice_cov),
        x0=pesi_iniziali,
        method="SLSQP",
        bounds=vincoli_pesi,
        constraints=lista_vincoli,
    )

    if not risultato.success:
        print(f"ATTENZIONE: ottimizzazione non perfettamente riuscita: {risultato.message}")

    return risultato


def trova_rendimento_massimo(rendimenti_attesi, vincoli_pesi):
    """
    Trova il rendimento massimo ottenibile rispettando i vincoli.
    Anche qui non blocchiamo tutto se scipy dà warning.
    """
    n_asset = len(rendimenti_attesi)
    pesi_iniziali = np.ones(n_asset) / n_asset

    risultato = minimize(
        fun=lambda w: -(w @ rendimenti_attesi),
        x0=pesi_iniziali,
        method="SLSQP",
        bounds=vincoli_pesi,
        constraints={"type": "eq", "fun": lambda w: np.sum(w) - 1},
    )

    if not risultato.success:
        print(f"ATTENZIONE: ottimizzazione rendimento massimo non perfettamente riuscita: {risultato.message}")

    return -risultato.fun


def genera_frontiera_efficiente(rendimenti_attesi, matrice_cov, vincoli_pesi, n_portafogli=10):
    """
    Genera n_portafogli sulla frontiera efficiente.
    """
    risultato_min_var = minimizza_varianza(matrice_cov, vincoli_pesi)
    rendimento_al_min_varianza = risultato_min_var.x @ rendimenti_attesi

    rendimento_al_max_varianza = trova_rendimento_massimo(rendimenti_attesi, vincoli_pesi)

    rendimenti_target = np.linspace(
        rendimento_al_min_varianza,
        rendimento_al_max_varianza,
        n_portafogli
    )

    portafogli_frontiera = []

    for rendimento_target in rendimenti_target:
        risultato = minimizza_varianza(
            matrice_cov,
            vincoli_pesi,
            rendimento_target=rendimento_target,
            rendimenti_attesi=rendimenti_attesi,
        )

        pesi_ottimali = risultato.x
        rendimento_realizzato = pesi_ottimali @ rendimenti_attesi
        volatilita = np.sqrt(calcola_varianza_portafoglio(pesi_ottimali, matrice_cov))

        if volatilita == 0:
            sharpe = 0
        else:
            sharpe = rendimento_realizzato / volatilita

        portafogli_frontiera.append({
            "pesi": pesi_ottimali,
            "rendimento": rendimento_realizzato,
            "volatilita": volatilita,
            "sharpe": sharpe,
        })

    return portafogli_frontiera


def stampa_frontiera_efficiente(portafogli_frontiera, nomi_asset):
    n_portafogli = len(portafogli_frontiera)

    print("\n=== Frontiera Efficiente — da Minima a Massima Varianza ===")
    print(f"{'#':<4} {'Rendimento':>12} {'Volatilità':>12} {'Sharpe':>10}")
    print("-" * 42)

    for i, portafoglio in enumerate(portafogli_frontiera):
        print(
            f"{i+1:<4} "
            f"{portafoglio['rendimento']:>12.4%} "
            f"{portafoglio['volatilita']:>12.4%} "
            f"{portafoglio['sharpe']:>10.4f}"
        )

    print("\n=== Allocazioni per Portafoglio ===")
    intestazione = f"{'Asset':<40}" + "".join(f"  P{i+1:<5}" for i in range(n_portafogli))
    print(intestazione)
    print("-" * len(intestazione))

    for j, nome_asset in enumerate(nomi_asset):
        riga = f"{str(nome_asset):<40}" + "".join(
            f"  {p['pesi'][j]:>5.1%}" for p in portafogli_frontiera
        )
        print(riga)


# =============================================================================
# FASE 3 — SALVATAGGIO MARKOWITZ
# =============================================================================

def salva_allocazioni_su_excel(portafogli_frontiera, nomi_asset, percorso_file):
    """
    Sovrascrive l'area C3:L20 del foglio 'risultati'.
    Salva anche rendimento e volatilità in C36:L37.
    """
    workbook = openpyxl.load_workbook(percorso_file, keep_vba=True)

    if "risultati" not in workbook.sheetnames:
        raise ValueError("Il foglio 'risultati' non esiste nel file Excel.")

    foglio_risultati = workbook["risultati"]

    riga_inizio = 3
    colonna_inizio = 3  # C

    for indice_asset, _ in enumerate(nomi_asset):
        for indice_portafoglio, portafoglio in enumerate(portafogli_frontiera):
            riga = riga_inizio + indice_asset
            colonna = colonna_inizio + indice_portafoglio
            peso = portafoglio["pesi"][indice_asset]
            foglio_risultati.cell(row=riga, column=colonna, value=round(float(peso), 6))

    for i, portafoglio in enumerate(portafogli_frontiera):
        foglio_risultati.cell(row=36, column=3 + i, value=round(float(portafoglio["rendimento"]), 6))
        foglio_risultati.cell(row=37, column=3 + i, value=round(float(portafoglio["volatilita"]), 6))

    workbook.save(percorso_file)

    print(
        f"\nRisultati salvati in '{percorso_file}' — "
        f"foglio 'risultati', area C3:L20 + C36:L37"
    )


# =============================================================================
# FASE 4 — MICHAUD-INSPIRED RESAMPLING CON BLOCK BOOTSTRAP
# =============================================================================

def _estrai_rendimenti_np(rendimenti_storici_df):
    solo_num = rendimenti_storici_df.select_dtypes(include=[np.number])
    solo_num = solo_num.dropna(how="all").fillna(0.0)
    return solo_num.values


def genera_campione_block_bootstrap(rendimenti_np, block_size=6):
    n_periodi = rendimenti_np.shape[0]

    if n_periodi == 0:
        raise ValueError("Non ci sono osservazioni storiche disponibili per il bootstrap.")

    block_size = min(block_size, n_periodi)

    blocchi = []
    n_accumulati = 0

    while n_accumulati < n_periodi:
        start = np.random.randint(0, n_periodi - block_size + 1)
        blocchi.append(rendimenti_np[start:start + block_size])
        n_accumulati += block_size

    return np.vstack(blocchi)[:n_periodi]


def _cov_da_campione_bootstrap(campione_np):
    cov = np.cov(campione_np, rowvar=False)

    if np.linalg.eigvalsh(cov).min() < 1e-10:
        cov += np.eye(campione_np.shape[1]) * 1e-8

    return cov


def ottimizzazione_michaud_resampling(
    rendimenti_attesi,
    rendimenti_storici_df,
    vincoli_pesi,
    n_portafogli=10,
    n_simulazioni=500,
    block_size=50,
):
    """
    Michaud-inspired resampling con block bootstrap.
    Mantiene la logica del tuo codice originale:
    se una simulazione fallisce, viene saltata.
    """
    rendimenti_np = _estrai_rendimenti_np(rendimenti_storici_df)

    if rendimenti_np.shape[0] == 0:
        raise ValueError("Il foglio 'Elaborazione' non contiene dati numerici validi.")

    pesi_accumulati = []

    std_rendimenti = rendimenti_np.std(axis=0)

    for h in range(n_simulazioni):
        try:
            campione = genera_campione_block_bootstrap(rendimenti_np, block_size)
            cov_bootstrap = _cov_da_campione_bootstrap(campione)

            rumore = np.random.normal(0, std_rendimenti)
            rendimenti_perturbati = rendimenti_attesi + rumore

            portafogli_h = genera_frontiera_efficiente(
                rendimenti_perturbati,
                cov_bootstrap,
                vincoli_pesi,
                n_portafogli=n_portafogli,
            )

            pesi_accumulati.append(np.array([p["pesi"] for p in portafogli_h]))

        except Exception as e:
            print(f"Simulazione {h + 1} fallita: {e}")

        if (h + 1) % 50 == 0:
            print(f"  Simulazione {h + 1}/{n_simulazioni} — riuscite: {len(pesi_accumulati)}")

    n_riuscite = len(pesi_accumulati)
    print(f"\nSimulazioni completate: {n_riuscite}/{n_simulazioni}")

    if n_riuscite == 0:
        raise RuntimeError("Nessuna simulazione Michaud è riuscita. Controlla dati, vincoli e rendimenti.")

    pesi_medi = np.stack(pesi_accumulati, axis=0).mean(axis=0)

    pesi_medi /= pesi_medi.sum(axis=1, keepdims=True)

    cov_originale = calcola_matrice_covarianza(rendimenti_storici_df) * 252

    portafogli_michaud = []

    for k in range(n_portafogli):
        pesi = pesi_medi[k]
        rendimento = pesi @ rendimenti_attesi
        volatilita = np.sqrt(calcola_varianza_portafoglio(pesi, cov_originale))

        if volatilita == 0:
            sharpe = 0
        else:
            sharpe = rendimento / volatilita

        portafogli_michaud.append({
            "pesi": pesi,
            "rendimento": rendimento,
            "volatilita": volatilita,
            "sharpe": sharpe,
        })

    return portafogli_michaud


def stampa_frontiera_michaud(portafogli_michaud, nomi_asset):
    n_portafogli = len(portafogli_michaud)

    print("\n=== Frontiera Michaud — Resampling Block Bootstrap ===")
    print(f"{'#':<4} {'Rendimento':>12} {'Volatilità':>12} {'Sharpe':>10}")
    print("-" * 42)

    for i, p in enumerate(portafogli_michaud):
        print(
            f"{i+1:<4} "
            f"{p['rendimento']:>12.4%} "
            f"{p['volatilita']:>12.4%} "
            f"{p['sharpe']:>10.4f}"
        )

    print("\n=== Allocazioni Michaud per Portafoglio ===")
    intestazione = f"{'Asset':<40}" + "".join(f"  P{i+1:<5}" for i in range(n_portafogli))
    print(intestazione)
    print("-" * len(intestazione))

    for j, nome in enumerate(nomi_asset):
        riga = f"{str(nome):<40}" + "".join(
            f"  {p['pesi'][j]:>5.1%}" for p in portafogli_michaud
        )
        print(riga)


def salva_michaud_su_excel(portafogli_michaud, nomi_asset, percorso_file):
    """
    Sovrascrive l'area B3:K20 del foglio 'Michaud'.
    Salva anche rendimento e volatilità in B27:K28.
    """
    workbook = openpyxl.load_workbook(percorso_file, keep_vba=True)

    if "Michaud" not in workbook.sheetnames:
        raise ValueError("Il foglio 'Michaud' non esiste nel file Excel.")

    foglio = workbook["Michaud"]

    riga_inizio = 3
    colonna_inizio = 2  # B

    for indice_asset, _ in enumerate(nomi_asset):
        for indice_portafoglio, portafoglio in enumerate(portafogli_michaud):
            foglio.cell(
                row=riga_inizio + indice_asset,
                column=colonna_inizio + indice_portafoglio,
                value=round(float(portafoglio["pesi"][indice_asset]), 6),
            )

    for i, portafoglio in enumerate(portafogli_michaud):
        foglio.cell(row=27, column=2 + i, value=round(float(portafoglio["rendimento"]), 6))
        foglio.cell(row=28, column=2 + i, value=round(float(portafoglio["volatilita"]), 6))

    workbook.save(percorso_file)

    print(
        f"\nRisultati Michaud salvati in '{percorso_file}' — "
        f"foglio 'Michaud', area B3:K20 + B27:K28"
    )


# =============================================================================
# FUNZIONE PRINCIPALE USATA DALL'APP
# =============================================================================

def run_optimizer(excel_path):
    """
    Funzione principale chiamata dall'app grafica.
    Il file Excel arriva dal pulsante 'Sfoglia file Excel'.
    """

    print("\n" + "=" * 70)
    print("AVVIO OTTIMIZZATORE")
    print("=" * 70)
    print(f"File selezionato: {excel_path}")

    # -------------------------------------------------------------------------
    # FASE 1 — LETTURA DATI
    # -------------------------------------------------------------------------

    datiElaborazione = pd.read_excel(
        excel_path,
        sheet_name="Elaborazione"
    )

    datiInput = pd.read_excel(
        excel_path,
        sheet_name="Input",
        usecols="B,D",
        header=1,
        nrows=20,
    ).dropna(axis=1, how="all")

    datiInput = datiInput.loc[:, ~datiInput.columns.str.startswith("Unnamed")]
    datiInput = datiInput.iloc[2:]

    vincoli = pd.read_excel(
        excel_path,
        sheet_name="Input",
        usecols="N,O",
        header=1,
        nrows=20,
    ).dropna(axis=1, how="all")

    vincoli.columns = ["Min", "Max"]
    vincoli = vincoli.iloc[2:]

    print("\nDati input letti correttamente.")
    print(f"Numero asset letti: {len(datiInput)}")
    print(f"Numero vincoli letti: {len(vincoli)}")

    if len(datiInput) == 0:
        raise ValueError("Il foglio 'Input' non contiene asset validi nelle colonne B e D.")

    if len(vincoli) == 0:
        raise ValueError("Il foglio 'Input' non contiene vincoli validi nelle colonne N e O.")

    if len(datiInput) != len(vincoli):
        raise ValueError(
            f"Numero asset e numero vincoli non coincidono: "
            f"{len(datiInput)} asset, {len(vincoli)} vincoli."
        )

    # -------------------------------------------------------------------------
    # PREPARAZIONE INPUT
    # -------------------------------------------------------------------------

    rendimenti_attesi = datiInput.iloc[:, 1].values.astype(float)

    matrice_covarianza = calcola_matrice_covarianza(datiElaborazione) * 252

    vincoli_pesi = list(zip(
        vincoli["Min"].values.astype(float),
        vincoli["Max"].values.astype(float),
    ))

    nomi_asset = datiInput.iloc[:, 0].values

    # Diagnostica utile, ma non blocca nulla
    lower_bounds = np.array([v[0] for v in vincoli_pesi])
    upper_bounds = np.array([v[1] for v in vincoli_pesi])

    print("\n=== CONTROLLO VINCOLI ===")
    print(f"Somma minimi: {lower_bounds.sum():.6f}")
    print(f"Somma massimi: {upper_bounds.sum():.6f}")
    print(f"Min più piccolo: {lower_bounds.min():.6f}")
    print(f"Max più grande: {upper_bounds.max():.6f}")
    print("=========================\n")

    # -------------------------------------------------------------------------
    # FASE 2 — MARKOWITZ
    # -------------------------------------------------------------------------

    print("\n" + "=" * 70)
    print("AVVIO OTTIMIZZAZIONE MARKOWITZ")
    print("=" * 70)

    portafogli_frontiera = genera_frontiera_efficiente(
        rendimenti_attesi,
        matrice_covarianza,
        vincoli_pesi,
        n_portafogli=10,
    )

    stampa_frontiera_efficiente(portafogli_frontiera, nomi_asset)

    # -------------------------------------------------------------------------
    # FASE 3 — SALVATAGGIO MARKOWITZ
    # -------------------------------------------------------------------------

    salva_allocazioni_su_excel(
        portafogli_frontiera,
        nomi_asset,
        excel_path
    )

    # -------------------------------------------------------------------------
    # FASE 4 — MICHAUD RESAMPLING
    # -------------------------------------------------------------------------

    print("\n" + "=" * 70)
    print("AVVIO MICHAUD RESAMPLING")
    print("=" * 70)

    mat_diag = datiElaborazione.select_dtypes(include=[np.number]).dropna(how="all").fillna(0.0)

    print(
        f"[DIAGNOSTICA] Osservazioni storiche disponibili: "
        f"{mat_diag.shape[0]} righe × {mat_diag.shape[1]} asset"
    )

    if mat_diag.shape[0] > 0:
        cov1 = np.cov(genera_campione_block_bootstrap(mat_diag.values, 6), rowvar=False)
        cov2 = np.cov(genera_campione_block_bootstrap(mat_diag.values, 6), rowvar=False)
        print(
            f"[DIAGNOSTICA] Differenza media tra due covarianze bootstrap: "
            f"{np.abs(cov1 - cov2).mean():.6f}"
        )

    print("Avvio Michaud Resampling: 500 simulazioni")

    portafogli_michaud = ottimizzazione_michaud_resampling(
        rendimenti_attesi=rendimenti_attesi,
        rendimenti_storici_df=datiElaborazione,
        vincoli_pesi=vincoli_pesi,
        n_portafogli=10,
        n_simulazioni=500,
        block_size=50,
    )

    stampa_frontiera_michaud(portafogli_michaud, nomi_asset)

    salva_michaud_su_excel(
        portafogli_michaud,
        nomi_asset,
        excel_path
    )

    print("\n" + "=" * 70)
    print("OTTIMIZZAZIONE COMPLETATA")
    print("=" * 70)

    return "Ottimizzazione completata correttamente. Risultati salvati nel file Excel selezionato."
#Alessio Bisceglia